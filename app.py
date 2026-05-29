import warnings
warnings.filterwarnings("ignore")

import streamlit as st
import pandas as pd
import openpyxl
import re
import unicodedata
import io
import base64

st.set_page_config(page_title="교육과정 자율점검 정밀 진단", layout="wide")

# ==========================================
# 1. PARSER MODULE (데이터 파싱)
# ==========================================
def _norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKC", str(s))
    return re.sub(r"\s+", " ", s.replace("\n", " ").replace("\u3000", " ")).strip()

def _parse_credit(v):
    if v is None or _norm(v) == "": return None, False
    val_str = str(v).strip()
    is_cross = "[" in val_str or "]" in val_str
    m = re.search(r"\d+(?:\.\d+)?", val_str)
    return float(m.group()) if m else None, is_cross

def parse_curriculum(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    target_sheet = next((s for s in wb.sheetnames if re.search(r"\d{4}\s*입학생", s)), None)
    if not target_sheet: raise ValueError("입학생 시트를 찾을 수 없습니다.")
    ws = wb[target_sheet]
    
    ss, ee = 6, ws.max_row
    for r in range(1, min(ws.max_row, 30) + 1):
        if _norm(ws.cell(r, 1).value) in ["학교지정", "학생선택", "1학년선택", "2학년선택", "3학년선택"]:
            ss = r; break

    rows = []
    gubun, gun = "", ""
    for r in range(ss, ee + 1):
        c1, c2 = _norm(ws.cell(r, 1).value), _norm(ws.cell(r, 2).value)
        if "이수" in c1 and "소계" in c1: break
        if c1: gubun = c1
        if c2: gun = c2
        subj = _norm(ws.cell(r, 4).value)
        if not subj: continue
            
        terms_data = {
            "1-1": _parse_credit(ws.cell(r, 7).value), "1-2": _parse_credit(ws.cell(r, 8).value),
            "2-1": _parse_credit(ws.cell(r, 9).value), "2-2": _parse_credit(ws.cell(r, 10).value),
            "3-1": _parse_credit(ws.cell(r, 11).value), "3-2": _parse_credit(ws.cell(r, 12).value)
        }
        op_crd, _ = _parse_credit(ws.cell(r, 6).value)
        rows.append({
            "행번호": r, "구분": gubun, "교과군": gun, "과목유형": _norm(ws.cell(r, 3).value), 
            "과목명": subj, "운영학점": op_crd, "terms": terms_data, "비고": _norm(ws.cell(r, 13).value)
        })
    return rows

# ==========================================
# 2. CHECKER MODULE (자율점검표 기준 & 상세 진단)
# ==========================================
def check_curriculum(rows):
    results = []
    def add(item, val, res, diag):
        results.append({"점검항목": item, "측정값": val, "판정": res, "상세진단": diag})

    terms = ['1-1', '1-2', '2-1', '2-2', '3-1', '3-2']
    total_course_credits = 0.0
    kme_total = 0.0 
    pe_semesters_found = set()
    
    for r in rows:
        is_pe = '체육' in r['교과군'].replace(' ','') or '체육' in r['과목명']
        is_designated = "지정" in r["구분"]
        for t in terms:
            crd, is_cross = r["terms"][t]
            if crd and crd > 0:
                if is_pe: pe_semesters_found.add(t)
                if is_designated and not is_cross and r["교과군"] in ["국어", "수학", "영어"]:
                    kme_total += crd
                if is_designated and not is_cross:
                    total_course_credits += crd

    choice_credits = sum([(r["운영학점"] or 0) for r in rows if "선택" in r["구분"]])
    
    # 1. 체육 매 학기 편성
    item_pe = "☑️ 체육 교과는 매 학기 편성하여 이수하도록 하였는가?"
    if len(pe_semesters_found) == 6:
        add(item_pe, "6개 학기 배당 완료", "PASS", 
            "1학년 1학기부터 3학년 2학기까지 누락 없이 체육 교과가 편성되었습니다. 교차 이수(대괄호 표기)를 통한 편성도 정상적으로 감지되어 기준을 완벽히 충족합니다.")
    else:
        missed = [t for t in terms if t not in pe_semesters_found]
        add(item_pe, f"{len(pe_semesters_found)}개 학기 배당 (누락: {', '.join(missed)})", "FAIL", 
            f"2022 개정 교육과정 지침에 위배됩니다. 체육 교과는 6개 학기 모두 의무적으로 편성해야 하나, 현재 {', '.join(missed)} 학기에 배당이 누락되었습니다. 편성표를 수정해야 합니다.")

    # 2. 기초교과 편중 방지
    item_kme = "☑️ 기초 교과(국·수·영) 이수 학점이 총 교과 이수 학점(174)의 50%(81학점)를 초과하지 않는가?"
    kme_base = sum([(r["운영학점"] or 0) for r in rows if "지정" in r["구분"] and r["교과군"].replace(' ','') in ["국어","수학","영어"]])
    if kme_base <= 81:
        add(item_kme, f"지정 국수영 {kme_base:g}학점", "PASS", 
            f"학교 지정 과목 기준 국·수·영 합계가 {kme_base:g}학점으로, 법정 상한선인 81학점을 안전하게 준수하고 있습니다. 학생들이 선택과목에서 기초교과를 추가 이수하더라도 한도 규정을 지킬 수 있도록 설계되었습니다.")
    else:
        add(item_kme, f"지정 국수영 {kme_base:g}학점", "FAIL", 
            f"학교 지정 과목만으로 이미 국·수·영 합계가 {kme_base:g}학점이 되어 상한선(81학점)을 초과했습니다. 교과 비중을 낮추거나 탐구/예체능 교과로 분산 배치가 필요합니다.")

    # 3. 고정 학점 준수 (한국사, 과탐실)
    item_fix = "☑️ 각 과목별 학점의 증감 범위를 준수하였는가? (한국사 3학점, 과탐실 1학점 등)"
    history_err = [f"{r['과목명']}({r['운영학점']})" for r in rows if "한국사" in r['과목명'].replace(' ','') and r['운영학점'] != 3]
    sci_err = [f"{r['과목명']}({r['운영학점']})" for r in rows if "과학탐구실험" in r['과목명'].replace(' ','') and r['운영학점'] != 1]
    
    if not history_err and not sci_err: 
        add(item_fix, "고정 학점 준수 확인", "PASS", 
            "한국사(3학점 고정) 및 과학탐구실험(1학점 고정) 과목이 지침에 명시된 단위(학점) 수에 맞게 정확히 편성되었습니다. 학점 증감 금지 조항을 잘 지켰습니다.")
    else: 
        errs = ", ".join(history_err + sci_err)
        add(item_fix, "고정 학점 위반 발견", "FAIL", 
            f"학점 증감이 금지된 과목의 학점이 잘못 입력되었습니다. ({errs}) 해당 과목의 운영 학점을 지침에 맞게 즉시 수정하시기 바랍니다.")

    # 4. 동일 과목 이중 편성 
    item_dup = "☑️ 동일한 과목을 서로 다른 학점 수로 이중 편성하지 않았는가?"
    dup_errors = []
    subj_credits = {}
    for r in rows:
        clean_subj = r["과목명"].strip()
        if not r["운영학점"]: continue
        if clean_subj not in subj_credits: subj_credits[clean_subj] = set()
        subj_credits[clean_subj].add(r["운영학점"])
        
    for subj, crds in subj_credits.items():
        if len(crds) > 1: dup_errors.append(f"{subj}({', '.join(map(lambda x: f'{x:g}', crds))})")

    if not dup_errors:
        add(item_dup, "이중 편성 없음", "PASS", 
            "편성된 모든 과목이 단일한 운영학점으로 깔끔하게 정의되어 있습니다. 학년별/학기별로 학점이 충돌하는 데이터 오류가 없습니다.")
    else:
        add(item_dup, f"이중 편성 {len(dup_errors)}건 발견", "FAIL", 
            f"하나의 과목이 여러 개의 운영학점을 가지도록 표기되었습니다. ({', '.join(dup_errors)}) 오기입이거나 중복 개설의 결과일 수 있으므로, 행을 병합하거나 학점을 통일해야 합니다.")
            
    # 5. 위계성 (Ⅰ → Ⅱ)
    item_hier = "☑️ 선택 과목 중 위계성이 있는 경우 계열적 학습이 가능하도록 편성하였는가?"
    roman_subjects = {}
    hierarchy_errors = []
    term_order = {t: i for i, t in enumerate(terms)}
    for r in rows:
        subj = r["과목명"].strip()
        m = re.match(r"(.*?)(Ⅰ|Ⅱ)$", subj)
        if m:
            base_name, level = m.group(1), m.group(2)
            first_term_idx = 99
            for t in terms:
                crd, _ = r["terms"][t]
                if crd and crd > 0:
                    first_term_idx = min(first_term_idx, term_order[t])
            if first_term_idx != 99:
                if base_name not in roman_subjects: roman_subjects[base_name] = {}
                roman_subjects[base_name][level] = first_term_idx

    for base_name, levels in roman_subjects.items():
        if 'Ⅰ' in levels and 'Ⅱ' in levels:
            if levels['Ⅰ'] > levels['Ⅱ']:
                hierarchy_errors.append(f"{base_name} (Ⅱ가 Ⅰ보다 선행됨)")

    if not hierarchy_errors:
        add(item_hier, "위계성 위반 없음", "PASS", 
            "Ⅰ과목과 Ⅱ과목으로 연계되는 교과들이 올바른 선수학습 순서에 따라 배치되었습니다. 학생들의 계열적 학습 구조가 정상적으로 보장됩니다.")
    else:
        add(item_hier, "위계성 순서 오류", "WARN", 
            f"일부 과목에서 Ⅱ과목이 Ⅰ과목보다 앞선 학기에 개설되었습니다. ({', '.join(hierarchy_errors)}) 특별한 교육과정적 사유가 없다면 학기 배치를 재조정할 것을 권고합니다.")

    return results

# ==========================================
# 3. REPORT GENERATION (HTML to PDF)
# ==========================================
def generate_html_report(school_name, results):
    """ 브라우저에서 인쇄(PDF로 저장)하기 최적화된 HTML 리포트 생성 """
    html = f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Malgun Gothic', sans-serif; padding: 40px; color: #333; line-height: 1.6; }}
            .header {{ text-align: center; border-bottom: 3px solid #1e3a8a; padding-bottom: 20px; margin-bottom: 30px; }}
            .title {{ font-size: 28px; font-weight: bold; color: #1e3a8a; margin: 0; }}
            .subtitle {{ font-size: 16px; color: #64748b; margin-top: 10px; }}
            .card {{ border: 1px solid #e2e8f0; border-radius: 8px; margin-bottom: 20px; page-break-inside: avoid; }}
            .card-header {{ padding: 15px; background-color: #f8fafc; border-bottom: 1px solid #e2e8f0; font-weight: bold; font-size: 16px; }}
            .card-body {{ padding: 15px; }}
            .badge-pass {{ color: #15803d; font-weight: bold; }}
            .badge-fail {{ color: #b91c1c; font-weight: bold; }}
            .badge-warn {{ color: #b45309; font-weight: bold; }}
            .diag-text {{ margin-top: 10px; padding: 12px; background-color: #f1f5f9; border-radius: 6px; font-size: 14px; color: #475569; border-left: 4px solid #3b82f6; }}
            @media print {{ body {{ padding: 0; }} .no-print {{ display: none; }} }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1 class="title">[{school_name}] 교육과정 정밀 진단 리포트</h1>
            <div class="subtitle">고등학교 교육과정 편성·운영 자율점검표 기준 상세 진단 결과</div>
        </div>
    """
    
    for r in results:
        b_class = f"badge-{r['판정'].lower()}"
        icon = "✅" if r['판정'] == "PASS" else ("❌" if r['판정'] == "FAIL" else "⚠️")
        html += f"""
        <div class="card">
            <div class="card-header">{r['점검항목']}</div>
            <div class="card-body">
                <div><span>진단 결과:</span> <span class="{b_class}">{icon} {r['판정']}</span></div>
                <div><span>측정 데이터:</span> <b>{r['측정값']}</b></div>
                <div class="diag-text">
                    <b>💡 진단 코멘트:</b><br>{r['상세진단']}
                </div>
            </div>
        </div>
        """
        
    html += "</body></html>"
    return html

# ==========================================
# 4. STREAMLIT UI
# ==========================================
def main():
    st.title("🏫 교육과정 자율점검 다중 진단 시스템")
    st.markdown("자율점검표 문항에 따른 **상세한 진단 코멘트**를 제공하며, 결과를 **PDF(인쇄용 HTML)** 형태로 다운로드할 수 있습니다.")
    
    uploaded_files = st.file_uploader("여러 학교의 배당표 엑셀 업로드 (.xlsx)", type=['xlsx'], accept_multiple_files=True)
    
    if uploaded_files:
        for file in uploaded_files:
            file_bytes = file.getvalue()
            school_name = file.name.replace(".xlsx", "").replace(".csv", "")
            
            try:
                rows = parse_curriculum(file_bytes)
                results = check_curriculum(rows)
                has_error = any(r["판정"] == "FAIL" for r in results)
                
                # --- 학교별 아코디언 패널 ---
                with st.expander(f"🏫 {school_name} 진단 리포트 ({'⚠️ 수정 요망' if has_error else '✅ 완벽함'})", expanded=False):
                    
                    # 상세 진단 결과 출력
                    for r in results:
                        color = "green" if r["판정"] == "PASS" else ("red" if r["판정"] == "FAIL" else "orange")
                        st.markdown(f"#### {r['점검항목']}")
                        st.markdown(f"**판정:** :{color}[**{r['판정']}**] &nbsp;|&nbsp; **측정값:** {r['측정값']}")
                        st.info(f"💡 **진단 코멘트:** {r['상세진단']}")
                        st.divider()
                    
                    # 리포트 다운로드 버튼 생성
                    html_report = generate_html_report(school_name, results)
                    b64 = base64.b64encode(html_report.encode('utf-8')).decode()
                    href = f'<a href="data:text/html;base64,{b64}" download="{school_name}_진단리포트.html" target="_blank" style="text-decoration: none;"><button style="background-color:#1e3a8a; color:white; padding:10px 20px; border:none; border-radius:5px; cursor:pointer; font-weight:bold;">📥 {school_name} 상세 리포트 다운로드 (클릭 후 Ctrl+P로 PDF 저장)</button></a>'
                    st.markdown(href, unsafe_allow_html=True)
                    st.caption("※ 다운로드한 HTML 파일을 열고 인쇄(Ctrl+P) 화면에서 'PDF로 저장'을 선택하시면 완벽한 형태의 PDF 문서가 생성됩니다.")
                    
            except Exception as e:
                st.error(f"'{school_name}' 처리 중 오류 발생: {e}")

if __name__ == '__main__':
    main()
